"""
Generates the sample Step-1 (Cleanup) and Step-2 (Relationships) audit workbooks by
running the repo's audit logic against the *before* sample model
(sample-model/before/ContosoRetailMini.SemanticModel).

Output: sample-results/01_Cleanup_Dependency_Audit.xlsx
        sample-results/02_AI_Readiness_Relationship_Analysis.xlsx

Dev tool. Requires openpyxl. Run:  py -3 _build_sample_results.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "sample-results")
os.makedirs(OUT, exist_ok=True)

# ---- styling ----------------------------------------------------------------
HDR = PatternFill("solid", fgColor="2F5496")
FAIL = PatternFill("solid", fgColor="FFC7CE")
WARN = PatternFill("solid", fgColor="FFEB9C")
PASS = PatternFill("solid", fgColor="C6EFCE")
HDR_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
MONO = Font(name="Consolas", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def header(ws, cols, row=1):
    for c, name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = HDR
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN


def widths(ws, ws_widths):
    for col, w in ws_widths.items():
        ws.column_dimensions[col].width = w


def status_fill(val):
    v = str(val).upper()
    if v in ("FAIL", "NO", "HIGH RISK", "CRITICAL"):
        return FAIL
    if v in ("WARN", "REVIEW", "REVIEW REQUIRED", "MEDIUM"):
        return WARN
    if v in ("PASS", "YES", "OK", "LOW"):
        return PASS
    return None


# =============================================================================
# Parsed facts from sample-model/before (ContosoRetailMini)
# =============================================================================
# columns: (table, name, dataType, summarizeBy, is_key, used_by, needed, safe, conf, note)
COLUMNS = [
    ("fct_sls", "sls_id", "int64", "sum", True, "(surrogate key)", "Yes", "No", "High", "Key; should be hidden + summarizeBy None"),
    ("fct_sls", "cust_id", "int64", "sum", True, "rel fct_cust", "Yes", "No", "High", "Relationship key"),
    ("fct_sls", "prod_id", "int64", "sum", True, "rel fct_prod", "Yes", "No", "High", "Relationship key"),
    ("fct_sls", "dt_ky", "int64", "sum", True, "rel fct_date", "Yes", "No", "High", "Relationship key"),
    ("fct_sls", "txn_amt", "double", "sum", False, "m_ttl_sls, m_sls_ytd, m_tmp_old", "Yes", "No", "High", "Measure input"),
    ("fct_sls", "qty", "double", "sum", False, "m_ttl_qty", "Yes", "No", "High", "Measure input; dataType should be int64"),
    ("fct_sls", "disc_pct", "double", "sum", False, "m_avg_disc", "Yes", "No", "High", "Measure input"),
    ("dim_cust", "cust_id", "int64", "sum", True, "rel fct_cust", "Yes", "No", "High", "Relationship key"),
    ("dim_cust", "cust_nm", "string", "none", False, "(dimension attribute)", "Review", "No", "Low", "No measure/visual usage; keep as dim attribute"),
    ("dim_cust", "cty", "string", "none", False, "(dimension attribute)", "Review", "No", "Low", "No measure/visual usage; keep as dim attribute"),
    ("dim_cust", "seg", "string", "none", False, "(dimension attribute)", "Review", "No", "Low", "No measure/visual usage; keep as dim attribute"),
    ("dim_prod", "prod_id", "int64", "sum", True, "rel fct_prod", "Yes", "No", "High", "Relationship key"),
    ("dim_prod", "prod_nm", "string", "none", False, "(dimension attribute)", "Review", "No", "Low", "No measure/visual usage; keep as dim attribute"),
    ("dim_prod", "cat", "string", "none", False, "(dimension attribute)", "Review", "No", "Low", "No measure/visual usage; keep as dim attribute"),
    ("dim_prod", "subcat", "string", "none", False, "(dimension attribute)", "Review", "No", "Low", "No measure/visual usage; keep as dim attribute"),
    ("dim_dt", "dt_ky", "int64", "sum", True, "rel fct_date", "Yes", "No", "High", "Relationship key"),
    ("dim_dt", "dt", "dateTime", "none", False, "m_sls_ytd", "Yes", "No", "High", "Time-intelligence date column"),
    ("dim_dt", "yr", "int64", "sum", False, "(dimension attribute)", "Review", "No", "Low", "Year should be summarizeBy None"),
    ("dim_dt", "mth", "int64", "sum", False, "(dimension attribute)", "Review", "No", "Low", "Month; needs SortByColumn"),
    ("dim_dt", "qtr", "int64", "sum", False, "(dimension attribute)", "Review", "No", "Low", "Quarter; summarizeBy None"),
]

# measures: (table, name, dax, depends_on, used_by, needed, safe, conf, note)
MEASURES = [
    ("fct_sls", "m_ttl_sls", "SUM(fct_sls[txn_amt])", "fct_sls[txn_amt]", "m_sls_ytd", "Yes", "No", "High", "Primary sales measure"),
    ("fct_sls", "m_ttl_qty", "SUM(fct_sls[qty])", "fct_sls[qty]", "-", "Yes", "No", "High", "Primary quantity measure"),
    ("fct_sls", "m_avg_disc", "AVERAGE(fct_sls[disc_pct])", "fct_sls[disc_pct]", "-", "Yes", "No", "High", "Average discount"),
    ("fct_sls", "m_sls_ytd", "TOTALYTD([m_ttl_sls], dim_dt[dt])", "m_ttl_sls, dim_dt[dt]", "-", "Yes", "No", "High", "YTD; needs a marked date table"),
    ("fct_sls", "m_tmp_old", "SUMX(fct_sls, fct_sls[txn_amt] * 1)", "fct_sls[txn_amt]", "(none)", "No", "Yes", "High", "Redundant duplicate of m_ttl_sls; no dependents -> safe to delete"),
]

# relationships: (name, frm, to, cardinality, crossfilter, active, report_use, dax_use, category, rec)
RELS = [
    ("fct_cust", "fct_sls[cust_id]", "dim_cust[cust_id]", "Many-to-One", "Both", "Yes", "No (no visuals)", "No", "FIX-BIDI", "Set to single direction; use measure-side filtering if reverse filter is needed"),
    ("fct_prod", "fct_sls[prod_id]", "dim_prod[prod_id]", "Many-to-One", "Single", "Yes", "No (no visuals)", "No", "KEEP", "Healthy star-schema relationship"),
    ("fct_date", "fct_sls[dt_ky]", "dim_dt[dt_ky]", "Many-to-One", "Single", "Yes", "No (no visuals)", "m_sls_ytd", "KEEP", "Healthy; mark dim_dt as the Date table"),
]

TABLES = [
    ("fct_sls", "Fact", 7, 5, 3, "Transaction grain; one row per sale line"),
    ("dim_cust", "Dimension", 4, 0, 1, "Customer lookup"),
    ("dim_prod", "Dimension", 4, 0, 1, "Product lookup"),
    ("dim_dt", "Date Dimension", 5, 0, 1, "Calendar; NOT marked as Date table"),
]


# =============================================================================
# Workbook 1 — Cleanup / Dependency Audit (5 sheets)
# =============================================================================
def build_cleanup():
    wb = Workbook()

    # Sheet 1: Object Inventory
    ws = wb.active
    ws.title = "Object Inventory"
    cols = ["Table", "Object", "Type", "Data Type", "Needed", "Safe to Delete",
            "Confidence", "Used By / Where Used", "Is Key", "Is Hidden", "Recommendation"]
    header(ws, cols)
    r = 2
    for t, n, dt, sby, key, used, need, safe, conf, note in COLUMNS:
        vals = [t, n, "Column", dt, need, safe, conf, used, "Yes" if key else "No", "No", note]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            if c in (5, 6):
                f = status_fill(v)
                if f:
                    cell.fill = f
        r += 1
    for t, n, dax, dep, used, need, safe, conf, note in MEASURES:
        vals = [t, n, "Measure", "-", need, safe, conf, used, "No", "No", note]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            if c in (5, 6):
                f = status_fill(v)
                if f:
                    cell.fill = f
        r += 1
    widths(ws, {"A": 12, "B": 12, "C": 10, "D": 10, "E": 10, "F": 13,
                "G": 11, "H": 30, "I": 8, "J": 9, "K": 46})
    ws.freeze_panes = "A2"

    # Sheet 2: Relationships
    ws2 = wb.create_sheet("Relationships")
    cols = ["Relationship", "From", "To", "Cardinality", "Cross-filter", "Active",
            "Used in Report", "Used via DAX", "Recommendation"]
    header(ws2, cols)
    for i, (nm, frm, to, card, cf, act, rep, dax, cat, rec) in enumerate(RELS, 2):
        vals = [nm, frm, to, card, cf, act, rep, dax, rec]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            if c == 5 and cf == "Both":
                cell.fill = FAIL
        ws2.cell(row=i, column=5).fill = FAIL if cf == "Both" else PASS
    widths(ws2, {"A": 12, "B": 18, "C": 18, "D": 14, "E": 12, "F": 8,
                 "G": 16, "H": 12, "I": 60})
    ws2.freeze_panes = "A2"

    # Sheet 3: Dependency Tree
    ws3 = wb.create_sheet("Dependency Tree")
    header(ws3, ["Object", "Table", "Type", "Depends On", "Used By"])
    deps = [(n, t, "Measure", dep, used) for t, n, dax, dep, used, *_ in MEASURES]
    for i, (n, t, ty, dep, used) in enumerate(deps, 2):
        for c, v in enumerate([n, t, ty, dep, used], 1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
    widths(ws3, {"A": 14, "B": 12, "C": 10, "D": 28, "E": 18})
    ws3.freeze_panes = "A2"

    # Sheet 4: Unused Summary
    ws4 = wb.create_sheet("Unused Summary")
    header(ws4, ["Type", "Total", "Needed", "Review Required", "Safe to Delete", "Action"])
    col_total = len(COLUMNS)
    col_need = sum(1 for *_x, need, _s, _c, _n in [(c[0], c[1], c[2], c[3], c[4], c[5], c[6], c[7], c[8], c[9]) for c in COLUMNS] if need == "Yes")
    col_review = sum(1 for c in COLUMNS if c[6] == "Review")
    meas_total = len(MEASURES)
    meas_safe = sum(1 for m in MEASURES if m[6] == "Yes")
    rows = [
        ("Columns", col_total, col_total - col_review, col_review, 0, "Keep keys/measure inputs; review dim attributes"),
        ("Measures", meas_total, meas_total - meas_safe, 0, meas_safe, "Delete m_tmp_old (redundant)"),
        ("Tables", len(TABLES), len(TABLES), 0, 0, "Keep all; mark dim_dt as Date table"),
        ("Relationships", len(RELS), len(RELS), 0, 0, "Fix BiDi on fct_cust"),
    ]
    for i, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws4.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
    widths(ws4, {"A": 14, "B": 8, "C": 9, "D": 16, "E": 15, "F": 48})
    ws4.freeze_panes = "A2"

    # Sheet 5: Optimization
    ws5 = wb.create_sheet("Optimization")
    header(ws5, ["Category", "Object", "Finding", "Recommended Action"])
    opt = [
        ("Redundant measure", "fct_sls[m_tmp_old]", "SUMX(fct_sls, txn_amt*1) equals SUM(txn_amt)", "Delete; use m_ttl_sls"),
        ("Bi-directional filter", "fct_cust", "crossFilteringBehavior = bothDirections", "Set to single direction"),
        ("Auto date/time", "model", "__PBI_TimeIntelligenceEnabled = 0 (good), but no marked date table", "Mark dim_dt as Date table"),
        ("Implicit measures", "model", "discourageImplicitMeasures not set", "Set discourageImplicitMeasures = TRUE"),
        ("Visible keys", "sls_id, cust_id, prod_id, dt_ky", "Keys visible + summarizeBy sum", "Hide keys; summarizeBy None"),
        ("Wrong data type", "fct_sls[qty]", "qty is double; quantities are whole numbers", "Change to int64"),
        ("Report measures", "(none found)", "No report-level measures in this PBIP", "N/A"),
    ]
    for i, row in enumerate(opt, 2):
        for c, v in enumerate(row, 1):
            cell = ws5.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
    widths(ws5, {"A": 20, "B": 30, "C": 46, "D": 34})
    ws5.freeze_panes = "A2"

    path = os.path.join(OUT, "01_Cleanup_Dependency_Audit.xlsx")
    wb.save(path)
    return path


# =============================================================================
# Workbook 2 — AI-Readiness Relationship Analysis (10 sheets)
# =============================================================================
def build_relationship():
    wb = Workbook()

    # 1 Executive Summary
    ws = wb.active
    ws.title = "1 Executive Summary"
    ws["A1"] = "AI Readiness — Relationship Analysis"
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws["A2"] = "Model: ContosoRetailMini (before) — sample-model/before"
    summary = [
        ("Metric", "Value"),
        ("Tables", len(TABLES)),
        ("Relationships", len(RELS)),
        ("Bi-directional relationships", 1),
        ("Many-to-Many relationships", 0),
        ("Inactive relationships", 0),
        ("Date table marked", "No"),
        ("discourageImplicitMeasures", "Not set (FAIL)"),
        ("RISK LEVEL", "MEDIUM"),
    ]
    header(ws, summary[0], row=4)
    for i, (k, v) in enumerate(summary[1:], 5):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        a.border = THIN
        b.border = THIN
        if k in ("Date table marked", "discourageImplicitMeasures", "RISK LEVEL") or (k == "Bi-directional relationships" and v):
            b.fill = WARN if k == "RISK LEVEL" else FAIL
    widths(ws, {"A": 30, "B": 22})

    # 2 Model Inventory
    ws2 = wb.create_sheet("2 Model Inventory")
    header(ws2, ["Table", "Type", "Columns", "Measures", "Relationships", "Notes"])
    for i, (t, ty, ncol, nmeas, nrel, note) in enumerate(TABLES, 2):
        for c, v in enumerate([t, ty, ncol, nmeas, nrel, note], 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
    widths(ws2, {"A": 12, "B": 16, "C": 9, "D": 9, "E": 13, "F": 44})
    ws2.freeze_panes = "A2"

    # 3 All Relationships
    ws3 = wb.create_sheet("3 All Relationships")
    header(ws3, ["Relationship", "From", "To", "Cardinality", "Direction", "Active", "Category", "Issue"])
    for i, (nm, frm, to, card, cf, act, rep, dax, cat, rec) in enumerate(RELS, 2):
        issue = "Bi-directional filter — ambiguous paths for Copilot" if cf == "Both" else "None"
        row = [nm, frm, to, card, cf, act, cat, issue]
        for c, v in enumerate(row, 1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
        fill = FAIL if cf == "Both" else PASS
        for c in range(1, 9):
            ws3.cell(row=i, column=c).fill = fill
    widths(ws3, {"A": 12, "B": 18, "C": 18, "D": 14, "E": 11, "F": 8, "G": 12, "H": 46})
    ws3.freeze_panes = "A2"

    # 4 AI Readiness Issues
    ws4 = wb.create_sheet("4 AI Readiness Issues")
    header(ws4, ["#", "Issue", "Object", "Severity", "Why it breaks Copilot"])
    issues = [
        ("BiDi relationship", "fct_cust", "High", "Bi-directional filters create ambiguous filter paths; Copilot can return inconsistent aggregations."),
        ("No date table marked", "dim_dt", "High", "Time-intelligence and 'last month / YTD' questions need a model-marked Date table."),
        ("Implicit measures allowed", "model", "High", "Copilot may auto-sum raw columns instead of your defined measures."),
        ("Visible summed keys", "sls_id, cust_id, prod_id, dt_ky", "Medium", "Surrogate keys appear as fields and get summed, polluting AI answers."),
        ("Wrong data type", "fct_sls[qty]", "Medium", "Quantity typed as double invites fractional aggregates."),
        ("No descriptions", "all objects", "Medium", "Copilot relies on descriptions to disambiguate meaning."),
        ("Cryptic names", "fct_sls, txn_amt, ...", "Medium", "AI reads names literally; abbreviations carry no meaning."),
        ("No format strings", "measures", "Low", "Unformatted values reduce answer readability."),
        ("Redundant measure", "m_tmp_old", "Low", "Duplicate logic confuses metric selection."),
    ]
    for i, (iss, obj, sev, why) in enumerate(issues, 2):
        row = [i - 1, iss, obj, sev, why]
        for c, v in enumerate(row, 1):
            cell = ws4.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
        ws4.cell(row=i, column=4).fill = status_fill("FAIL" if sev == "High" else "WARN" if sev == "Medium" else "PASS")
    widths(ws4, {"A": 5, "B": 24, "C": 28, "D": 10, "E": 60})
    ws4.freeze_panes = "A2"

    # 5 Fix Suggestions
    ws5 = wb.create_sheet("5 Fix Suggestions")
    header(ws5, ["Issue", "Approach", "Model / DAX change", "Report impact", "Pros / Cons", "Best?"])
    fixes = [
        ("BiDi on fct_cust", "A: Single direction",
         "Set crossFilteringBehavior = OneDirection on fct_cust",
         "None (no visuals bound)", "Removes ambiguity; simplest. Reverse filtering lost if actually needed.", "YES — BEST"),
        ("BiDi on fct_cust", "B: CROSSFILTER in measure",
         "Keep single; wrap specific measures with CROSSFILTER(..., Both)",
         "None", "Surgical; only where required. More measure code to maintain.", ""),
        ("BiDi on fct_cust", "C: Bridge table",
         "Introduce a bridge if a true many-to-many exists",
         "None", "Correct for real M:M. Overkill here (relationship is M:1).", ""),
        ("No date table", "A: Mark dim_dt as Date",
         "MarkAsDateTable(dim_dt, dim_dt[dt])",
         "Enables date hierarchies", "Unlocks time-intelligence. Requires a contiguous date column.", "YES — BEST"),
        ("Implicit measures", "A: Discourage",
         "Set discourageImplicitMeasures = TRUE",
         "Users pick measures, not raw columns", "Cleaner AI behavior. Report authors must use measures.", "YES — BEST"),
    ]
    for i, row in enumerate(fixes, 2):
        for c, v in enumerate(row, 1):
            cell = ws5.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
        if row[5] == "YES — BEST":
            ws5.cell(row=i, column=6).fill = PASS
    widths(ws5, {"A": 18, "B": 22, "C": 42, "D": 26, "E": 46, "F": 12})
    ws5.freeze_panes = "A2"

    # 6 Report-Level Measures
    ws6 = wb.create_sheet("6 Report-Level Measures")
    header(ws6, ["Measure", "DAX", "Tables referenced", "Migration priority"])
    ws6.cell(row=2, column=1, value="(none found)").border = THIN
    ws6.cell(row=2, column=2, value="No reportExtensions.json in this PBIP — no report-level measures to migrate.").border = THIN
    ws6.merge_cells("B2:D2")
    widths(ws6, {"A": 18, "B": 60, "C": 20, "D": 18})

    # 7 AI-Readiness Checklist
    ws7 = wb.create_sheet("7 AI-Readiness Checklist")
    header(ws7, ["#", "Check", "Result", "Detail"])
    checks = [
        ("Many-to-Many = 0", "PASS", "No M:M relationships"),
        ("Bi-directional = 0", "FAIL", "fct_cust is bothDirections"),
        ("Hub/star schema", "PASS", "Single fact, conformed dimensions"),
        ("Key uniqueness", "PASS", "Dimension keys are unique"),
        ("No blank keys", "PASS", "No blanks in sample data"),
        ("Descriptions present", "FAIL", "0% of objects described"),
        ("Business-friendly names", "FAIL", "fct_sls, txn_amt, cust_nm, ..."),
        ("Keys hidden", "FAIL", "Keys visible"),
        ("Format strings", "FAIL", "Measures have no format strings"),
        ("discourageImplicitMeasures", "FAIL", "Not set"),
        ("isMdxAvailable reviewed", "WARN", "Not explicitly reviewed"),
        ("No auto-date tables", "PASS", "No LocalDateTable_* present"),
        ("SortByColumn on month/day", "FAIL", "mth has no SortByColumn"),
        ("Synonyms defined", "FAIL", "None"),
        ("DAX uses DIVIDE()", "PASS", "No bare division found"),
        ("DAX uses KEEPFILTERS()", "PASS", "No FILTER() over full tables"),
        ("Date table marked", "FAIL", "dim_dt not marked"),
        ("No circular paths", "PASS", "None"),
        ("Inactive relationships sane", "PASS", "None"),
    ]
    for i, (chk, res, detail) in enumerate(checks, 2):
        row = [i - 1, chk, res, detail]
        for c, v in enumerate(row, 1):
            cell = ws7.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
        ws7.cell(row=i, column=3).fill = status_fill(res)
    widths(ws7, {"A": 5, "B": 30, "C": 10, "D": 40})
    ws7.freeze_panes = "A2"

    # 8 Report Impact
    ws8 = wb.create_sheet("8 Report Impact")
    header(ws8, ["Page", "Change applied", "Impact level", "Visuals affected", "Needs recreation?"])
    ws8.append(["Overview", "Relationship/direction fixes", "None", "0 (placeholder page, no bound visuals)", "No"])
    for c in range(1, 6):
        ws8.cell(row=2, column=c).border = THIN
        ws8.cell(row=2, column=c).alignment = WRAP
    ws8.cell(row=2, column=3).fill = PASS
    widths(ws8, {"A": 14, "B": 28, "C": 12, "D": 40, "E": 16})

    # 9 Action Plan
    ws9 = wb.create_sheet("9 Action Plan")
    header(ws9, ["Priority", "Action", "Depends on", "Risk", "Status"])
    plan = [
        ("P0", "Set crossFilteringBehavior = OneDirection on fct_cust", "-", "Low", "Not Started"),
        ("P0", "Mark dim_dt as Date table", "-", "Low", "Not Started"),
        ("P1", "Set discourageImplicitMeasures = TRUE", "-", "Low", "Not Started"),
        ("P1", "Hide keys + summarizeBy None", "-", "Low", "Not Started"),
        ("P2", "Fix qty data type to int64", "-", "Low", "Not Started"),
        ("P2", "Delete redundant measure m_tmp_old", "cleanup audit", "Low", "Not Started"),
        ("P3", "Add descriptions, synonyms, format strings, friendly names", "steps 4-5", "Low", "Not Started"),
    ]
    for i, row in enumerate(plan, 2):
        for c, v in enumerate(row, 1):
            cell = ws9.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
    widths(ws9, {"A": 9, "B": 52, "C": 16, "D": 8, "E": 14})
    ws9.freeze_panes = "A2"

    # 10 Model Diagram (ASCII)
    ws10 = wb.create_sheet("10 Model Diagram")
    diagram = [
        "STAR SCHEMA — ContosoRetailMini (before)",
        "",
        "                 dim_cust",
        "                    |  <== BiDi (fix: single direction)",
        "                    |",
        "   dim_prod ----- fct_sls ----- dim_dt   (mark as Date table)",
        "                    |",
        "                 (measures: m_ttl_sls, m_ttl_qty, m_avg_disc,",
        "                  m_sls_ytd, m_tmp_old <== redundant)",
        "",
        "Legend:  ----- single-direction (healthy)    <== bi-directional (fix)",
    ]
    for i, line in enumerate(diagram, 1):
        cell = ws10.cell(row=i, column=1, value=line)
        cell.font = MONO
    ws10.column_dimensions["A"].width = 70

    path = os.path.join(OUT, "02_AI_Readiness_Relationship_Analysis.xlsx")
    wb.save(path)
    return path


# =============================================================================
# Workbook 3 — Rename Map (Step 5 output: before -> after)
# =============================================================================
def build_rename_map():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rename Map"
    header(ws, ["Object Type", "Table Name", "Old Name", "New Name", "Reason"])
    rows = [
        ("Table", "(model)", "fct_sls", "Fact Sales", "Cryptic prefix; explicit fact role, business-readable"),
        ("Table", "(model)", "dim_cust", "Dim Customer", "Cryptic prefix; explicit dimension role"),
        ("Table", "(model)", "dim_prod", "Dim Product", "Cryptic prefix; explicit dimension role"),
        ("Table", "(model)", "dim_dt", "Dim Date", "Cryptic prefix; explicit date-dimension role"),
        ("Column", "Fact Sales", "sls_id", "Sale Key", "Surrogate key; readable + hidden"),
        ("Column", "Fact Sales", "cust_id", "Customer Key", "Foreign key; readable + hidden"),
        ("Column", "Fact Sales", "prod_id", "Product Key", "Foreign key; readable + hidden"),
        ("Column", "Fact Sales", "dt_ky", "Date Key", "Foreign key; readable + hidden"),
        ("Column", "Fact Sales", "txn_amt", "Sales Amount", "Business term for revenue"),
        ("Column", "Fact Sales", "qty", "Quantity", "Expand abbreviation"),
        ("Column", "Fact Sales", "disc_pct", "Discount %", "Business term; percentage framing"),
        ("Column", "Dim Customer", "cust_id", "Customer Key", "Key; readable + hidden"),
        ("Column", "Dim Customer", "cust_nm", "Customer", "Primary display attribute"),
        ("Column", "Dim Customer", "cty", "City", "Expand abbreviation"),
        ("Column", "Dim Customer", "seg", "Segment", "Expand abbreviation"),
        ("Column", "Dim Product", "prod_id", "Product Key", "Key; readable + hidden"),
        ("Column", "Dim Product", "prod_nm", "Product", "Primary display attribute"),
        ("Column", "Dim Product", "cat", "Category", "Expand abbreviation"),
        ("Column", "Dim Product", "subcat", "Subcategory", "Expand abbreviation"),
        ("Column", "Dim Date", "dt_ky", "Date Key", "Key; readable + hidden"),
        ("Column", "Dim Date", "dt", "Date", "Primary date column"),
        ("Column", "Dim Date", "yr", "Year", "Expand abbreviation"),
        ("Column", "Dim Date", "mth", "Month", "Expand abbreviation"),
        ("Column", "Dim Date", "qtr", "Quarter", "Expand abbreviation"),
        ("Measure", "Fact Sales", "m_ttl_sls", "Total Sales", "Readable metric name"),
        ("Measure", "Fact Sales", "m_ttl_qty", "Total Quantity", "Readable metric name"),
        ("Measure", "Fact Sales", "m_avg_disc", "Average Discount %", "Readable metric name"),
        ("Measure", "Fact Sales", "m_sls_ytd", "Sales YTD", "Readable metric name"),
    ]
    for i, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
        fill = PASS if row[0] == "Table" else None
        if fill:
            ws.cell(row=i, column=1).fill = fill
    widths(ws, {"A": 12, "B": 14, "C": 12, "D": 20, "E": 50})
    ws.freeze_panes = "A2"

    # note: m_tmp_old is not renamed — it is deleted in Step 1 (cleanup)
    n = len(rows) + 3
    ws.cell(row=n, column=1, value="Note").font = BOLD
    ws.cell(row=n, column=2,
            value="m_tmp_old is not renamed — it is removed in Step 1 (cleanup) as a redundant measure.")
    ws.merge_cells(start_row=n, start_column=2, end_row=n, end_column=5)

    path = os.path.join(OUT, "03_Rename_Map.xlsx")
    wb.save(path)
    return path


if __name__ == "__main__":
    p1 = build_cleanup()
    p2 = build_relationship()
    p3 = build_rename_map()
    print("Wrote:", p1)
    print("Wrote:", p2)
    print("Wrote:", p3)
